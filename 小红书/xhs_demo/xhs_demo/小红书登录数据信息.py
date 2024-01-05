import os
import random
import re
from concurrent.futures import ThreadPoolExecutor
from hashlib import md5
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import datetime,xlwt,xlrd
from lxml import etree
import json
import time
from xlutils.copy import copy
import requests
from loguru import logger
import execjs

from xhs_login import get_session

class XHS:
    def __init__(self, web_session):
        self.sess = requests.session()
        self.x_s_keys = [187050025, 472920585, 186915882, 876157969, 255199502, 806945584, 220596020, 958210835,
                         757275681, 940378667, 489892883, 705504304, 354103316, 688857884, 890312192, 219096591,
                         622400037, 254088489, 907618332, 52759587, 907877143, 53870614, 839463457, 389417746,
                         975774727, 372382245, 437136414, 909246726, 168694017, 473575703, 52697872,
                         1010440969]  # 生成 x-s 的des密钥
        self.profileData_keys = [187567141, 875696391, 170266120, 876222754, 188089115, 1010309137, 187054378,
                                 957950720, 758514978, 941162813, 221382708, 990709537, 758848528, 688730163, 890444313,
                                 722272792, 890962233, 252521496, 890843430, 185009704, 874317360, 119997734, 907612693,
                                 119932961, 841824786, 120993794, 839716879, 909248796, 439099654, 372901635, 439091750,
                                 1009915397]  # 生成 profileDate 的des密钥
        # self.category = category  # 获取首页分类标签

        # 加载加密函数
        with open('des.js', 'r', encoding='utf8') as f:
            jsCode = f.read()
        self.encrypt_func = execjs.compile(jsCode)
        # 需要加密参数
        self.a1 = self.encrypt_func.call('get_a1')
        self.profileData = self.encrypt_func.call("get_profiledata", self.profileData_keys)
        self.encrypt_url = ""  # 网址
        self.x_s = ""
        self.x_t = ""
        self.x_s_common = ""
        self.headers = {
            "authority": "edith.xiaohongshu.com",
            "content-type": "application/json;charset=UTF-8",
            "accept": "application/json, text/plain, */*",
            "accept-language": "zh-CN,zh;q=0.9",
            "cache-control": "no-cache",
            "origin": "https://www.xiaohongshu.com",
            "pragma": "no-cache",
            "referer": "https://www.xiaohongshu.com/",
            "sec-ch-ua": "\"Google Chrome\";v=\"117\", \"Not;A=Brand\";v=\"8\", \"Chromium\";v=\"117\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-site",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36",
            "x-b3-traceid": "31ff71930b4371d2",
        }
        self.sess.cookie = {

            # "webBuild": "2.11.5",
            "webId": "10091d8528b024c76e3b713f8b559d0b",
            "xsecappid": "xhs-pc-web",
            "a1": self.a1,
            'websectiga': 'cffd9dcea65962b05ab048ac76962acee933d26157113bb213105a116241fa6c',
            'sec_poison_id': 'bb94474b-7744-4c48-aa64-2a3fcf8b4a93',
            # 'acw_tc': '56536d0f629a0c348744a01ede90644e6185baa519f1dc91e1eaf683e48b8ea7',
            'gid': 'yYDJSDd8JdqqyYDJSDd88dJ7q0df8Ed2KIqCJ2V3I4D4MY28Y3uM8Y888qyJyjy8fjW4j8qj',
            'abRequestId': '3b73179c-b780-5e78-9ec4-40c6644d85ae',
            "web_session": web_session,
        }

        self.response = ""
        self.index = 0
        self.option = ['img_list', 'video_url']

    def post_url(self, url, data, x10, requestMethod: str = "POST"):

        logger.debug(f'访问 {url.split("xiaohongshu.com")[1]} start')
        self.encrypt_url = f'url={url.split("xiaohongshu.com")[1]}{data}'

        self.x_t = int(round(time.time() * 1000))
        url_toMd5Hex = self.encrypt_func.call('toMd5Hex', self.encrypt_url)

        encrypt_str = f'x1={url_toMd5Hex};x2=0|0|0|1|0|0|1|0|0|0|1|0|0|0|0;x3={self.a1};x4={self.x_t};'

        payload = self.encrypt_func.call('get_argus', self.x_s_keys, encrypt_str)

        self.x_s = self.encrypt_func.call('get_x_s', payload)

        self.x_s_common = self.encrypt_func.call('get_x_s_common', self.a1, self.x_t, self.x_s, x10)
        logger.info(f"获取x_s：{self.x_s}")
        logger.info(f"获取x_s_common：{self.x_s_common}")
        logger.info(f"获取x-t：{self.x_t}")
        self.headers["x-s"] = self.x_s
        self.headers["x-s-common"] = self.x_s_common
        self.headers["x-t"] = str(self.x_t)
        print(self.headers)
        print(self.sess.cookie)

        if requestMethod == "POST":
            self.response = self.sess.post(url, headers=self.headers, data=data, cookies=self.sess.cookie)
        else:
            self.response = self.sess.get(url, headers=self.headers, params=data, cookies=self.sess.cookie)

        logger.debug(f'{url.split("xiaohongshu.com")[1]} status :{self.response.status_code}')
        return

    # 首页分类数据
    def get_search_by_topic(self, category: str):
        """

        Args:
            category: data里面的 category参数 翻页

        Returns:

        """
        post_url = "https://edith.xiaohongshu.com/api/sns/web/v1/homefeed"
        post_data = json.dumps({
            "cursor_score": "",
            "num": 39,
            "refresh_type": 1,
            "note_index": 0,
            "unread_begin_note_id": "",
            "unread_end_note_id": "",
            "unread_note_count": 0,
            "category": category,
            "search_key": ""
        }).strip()
        x_s_common_x10 = 8
        self.post_url(url=post_url, data=post_data, x10=x_s_common_x10)
        # if self.response.status_code == 200:
        logger.success(f'{self.response.text}')

    # 搜索指定帖子数据
    def search_topic(self, keyword: str):
        """
            搜索指定帖子, note_typ：0 全部，1 视频，2 图文
            sort = ['general', 'popularity_descending', 'time_descending']  # 排序 默认，最热，最新

        :param keyword:
        :return:
        """
        post_url = "https://edith.xiaohongshu.com/api/sns/web/v1/search/notes"
        page = 1
        post_data = {
            "keyword": keyword,
            "page": page,
            "page_size": 20,
            "search_id": "2cb5nfon0pwvxuu0o6a7a",
            "sort": "general",
            "note_type": 0,  # 搜索指定帖子, note_typ：0 全部，1 视频，2 图文
            "image_scenes": "FD_PRV_WEBP,FD_WM_WEBP"
        }
        data = json.dumps(post_data, separators=(',', ':'))

        x_s_common_x10 = 0
        continuations = [data]
        while continuations:
            self.post_url(url=post_url, data=data, x10=x_s_common_x10)

            if self.response.status_code != 200:
                logger.debug(f'状登录状态失效')
                yield self.response  # 461遇到滑块,待定后续修改

            elif self.response.json()['code'] == -100:
                logger.error('登录状态失效')
                yield -1  # 登录过期

            has_more = self.response.json().get('data', {}).get('has_more')

            # 存在下一页就翻页
            if has_more:
                page += 1
                post_data['page'] = page
                continuations.append(data)
            data_ = self.response.json().get('data', {}).get('items')

            if data_:
                yield from self.get_note(data_)
            time.sleep(random.uniform(1, 5))

    # 获取评论
    def get_note_comment(self, note_id: str):
        """

        Args:
            note_id: params 的note_id

        Returns:

        """
        x_s_common_x10 = 0
        continuations = ['']
        post_url = "https://edith.xiaohongshu.com/api/sns/web/v2/comment/page"

        while continuations:
            continuation = continuations.pop()

            params = {
                "note_id": note_id,
                "cursor": continuation,
                "top_comment_id": "",
                "image_scenes": "FD_WM_WEBP,CRD_WM_WEBP"
            }

            self.post_url(url=post_url, data=params, x10=8, requestMethod='GET')

            print(111111, self.response.json())
            has_more = self.response.json().get('data', {}).get('has_more')

            # 存在更多评论点击更多评论
            if has_more:
                continuations.append(self.response.json()['data']['cursor'])
            comments = self.response.json().get('data', {}).get('comments')

            if comments:
                yield from self.get_comment(note_id, comments)
            time.sleep(random.uniform(1, 5))
        return

    # 用户粉丝ip等详细信息获取
    def get_user(self, uid: str, retry_times=3):
        for _ in range(retry_times):
            post_url = f"https://www.xiaohongshu.com/user/profile/{uid}"

            self.post_url(url=post_url, data=[], x10=8, requestMethod='GIT')
            tree = etree.HTML(self.response.text)
            xingbie = re.search(r'xlink:href\s*=\s*"(#male|#female)"', self.response.text)
            if xingbie:
                if xingbie.group(1) == '#male':
                    xb = '男'
                else:
                    xb = '女'
            else:
                xb = '未知'
            aaa = tree.xpath('//*[@id="userPageContainer"]/div[1]/div/div[2]/div[1]')
            for i in aaa:
                # user_name = i.xpath('./div/div[2]/div[1]/div/text()')
                # if user_name:
                #     user_name = user_name[0]
                user_redId = i.xpath('./div/div[2]/div[2]/span[1]/text()')
                if not user_redId:
                    break
                try:
                    user_IP = i.xpath('./div/div[2]/div[2]/span[2]/text()')[0]
                    user_IP1 = user_IP.split('：')[-1]
                except:
                    user_IP1 = '未知'

                # user_desc = i.xpath('./div[2]/text()')[0]
                #
                # user_tags = i.xpath('./div[3]//text()')
                # if user_tags:
                #     user_desc = user_desc[0]
                user_interactions = i.xpath('./div[4]//text()')
                try:
                    user_interactions[4]
                except:
                    user_interactions = i.xpath('./div[3]//text()')

                try:
                    user_interactions[4]
                except:
                    user_interactions = i.xpath('./div[2]//text()')

                # if '万' in
                # item = {
                #     'user_name': user_name,
                #     'user_redId': user_redId[0],
                #     'user_IP': user_IP,
                #     'user_desc': user_desc,
                #     'user_tags': user_tags,
                #     'user_interactions': user_interactions,
                # }

            return xb,user_IP1,user_interactions[0],user_interactions[2],user_interactions[4]

    # 搜索用户信息
    def get_usersearch(self, keyword: str):
        page = 1
        post_url = "https://edith.xiaohongshu.com/api/sns/web/v1/search/usersearch"
        data = {
            "search_user_request": {
                "keyword": keyword,
                "search_id": "2cb6pcdrmucnzqixeiwkp",
                "page": page,
                "page_size": 15,
                "biz_type": "web_search_user",
                "request_id": "206220919-1697112504684"
            }
        }
        post_data = json.dumps(data, separators=(',', ':'))

        x_s_common_x10 = 0
        continuations = [data]
        while continuations:
            self.post_url(url=post_url, data=post_data, x10=x_s_common_x10)
            print(self.response)
            print(self.response.text)
            has_more = self.response.json().get('data', {}).get('has_more')
            # 存在下一页就翻页
            if has_more:
                page += 1
                data['page'] = page
                continuations.append(post_data)
            data_ = self.response.json().get('data', {}).get('users')
            print(data_)
            if data_:
                yield from self.get_search_user(data_)
            time.sleep(random.uniform(1, 5))

    # 获取帖子中的tag_list，这个就是详情页里面的话题
    def get_note_topic(self, note_id: str):
        post_url = f'https://www.xiaohongshu.com/fe_api/burdock/v2/note/{note_id}/tags'
        # 标准的md5算法
        obj = md5()
        obj.update(f'/fe_api/burdock/v2/note/{note_id}/tagsWSUDD'.encode('utf-8'))
        self.headers['x-sign'] = 'X' + obj.hexdigest()

        self.post_url(url=post_url, data=[], x10=8, requestMethod='GET')
        try:
            data = [{
                'name': topic['name'],
                'link': topic['link'],
                'id': topic['pageId']
            } for topic in self.response.json()['data']]
            return data
        except KeyError:
            raise KeyError(self.response.json()['msg'])

    # 定义一个函数，用于判断字符串是否包含 "mp4" 并提取链接
    def filter_mp4(self, string: str):
        if isinstance(string, dict):
            for value in string.values():
                for link in value:
                    if "mp4" in link:
                        return link
        elif isinstance(string, str):
            if "mp4" in string:
                return string
        return None

    # 获取笔记的详细信息
    def get_note_detail(self, note_id: str):
        post_url = 'https://edith.xiaohongshu.com/api/sns/web/v1/feed'
        data = {
            'source_note_id': note_id,
            "image_scenes": [
                "CRD_PRV_WEBP",
                "CRD_WM_WEBP"
            ],
            "extra": {
                "need_body_topic": "1"
            }

        }
        post_data = json.dumps(data, separators=(',', ':'))
        self.post_url(url=post_url, data=post_data, x10=0)
        data = self.response.json().get('data').get('items')
        if data:
            d = data[0]
            # model_type = d.get('model_type')
            note_id = d.get('id')
            i = d.get('note_card')
            title = i.get('title')
            note_type = i.get('type')
            desc = i.get('desc')
            # at_user_list = i.get('at_user_list')
            nickname = i.get('user', {}).get('nickname')
            user_id=i.get('user', {}).get('user_id')
            # video = i.get('video', {}).get('media', {}).get('stream', {})
            # if video:
            #     video_url = [v.get('master_url') for v in video.get('h264', {})]
            # else:
            #     video_url = None
            # img_list = [pic.get('info_list')[0]['url'] for pic in i.get('image_list', {})]
            # tag_list = self.get_note_topic(note_id)
            count = i.get('interact_info')
            time_1 = i.get('time')
            timestamp_s = time_1 / 1000
            dt_object = datetime.datetime.fromtimestamp(timestamp_s)
            formatted_date = dt_object.strftime("%Y-%m-%d %H:%M:%S")
            like_count = count.get('liked_count') if i.get('interact_info') else None
            share_count = count.get('share_count') if i.get('interact_info') else None
            collected_count = count.get('collected_count') if i.get('interact_info') else None
            comment_count = count.get('comment_count') if i.get('interact_info') else None

            # item = {
            #     'model_type': note_type,
            #     'note_id': note_id,
            #     'title': title,
            #     'desc': desc,
            #     'user_name': nickname,
            #     'video_url': video_url,
            #     'img_list': img_list,
            #     'tag_list': tag_list,
            #     'at_user_list': at_user_list,
            #     'like_count': like_count,
            #     'share_count': share_count,
            #     'collected_count': collected_count,
            #     'comment_count': comment_count
            # }
            xb,ipsd,gzs,fss, dzyss= self.get_user(user_id)
            # print(k)
            lianjie = 'https://www.xiaohongshu.com/explore/'+note_id
            # item = {
            #     'lianjie': 'https://www.xiaohongshu.com/explore/'+note_id,
            #     'zuozhe': nickname,
            #     'yonghuid':user_id,
            #     'biaoti':title,
            #     'neirong':desc,
            #     'dianzhanshu':like_count,
            #     'shouchangshu':collected_count,
            #     'fenxiangshu':share_count,
            #     'pinglun':comment_count,
            #     'bijileixing':note_type,
            #     'xb':xb,
            #     'ipsd':ipsd,
            #     'gzs':gzs,
            #     'fss':fss,
            #     "dzyss":dzyss,
            #     'sj':formatted_date
            #
            # }

            data = {
                '数据':[lianjie,nickname,user_id,title,desc,like_count,collected_count,share_count,comment_count,note_type,xb,ipsd,gzs,fss,dzyss,formatted_date]
            }
            self.save_data(data)

            return data

    # 获取需要下载的全部链接 , 这里入库的时候使用
    def get_all_urls(self, note_id: str):
        url_list = []

        rep = self.get_note_detail(note_id)

        if rep['video_url']:
            url_list.append(rep['video_url'][0])
        else:
            for i in rep['img_list']:
                url_list.append(i)
        return url_list

    # 下载 本地使用.入库入下载链接就行了
    @staticmethod
    def download(pic, option: str, retry_times: int = 5):
        if not os.path.exists('./Lib'):
            os.mkdir('./Lib')
        flag = ['.jpg', '.mp4']
        for _ in range(retry_times):

            try:
                path = re.sub(r'[^a-zA-Z0-9]+', '', pic.split('/')[-1])[:38] + flag[option]

                data = requests.get(pic, timeout=10)
                with open('./Lib/' + path, 'wb') as fp:
                    fp.write(data.content)
                print(pic)
                return True
            except Exception as e:
                print(f"Error:{e}, retry_times: {_ + 1}/{retry_times}...")
                time.sleep(random.uniform(1, 5))
        print(f"{pic} download Failed...")

        # 获取用户主页的帖子和首页的帖子

    @staticmethod
    def get_note(notes: str):
        for note in notes:
            id = note.get('note_id')
            if note.get('hot_query'):
                # 推荐热门搜搜
                continue
            if note.get('note_card'):
                id = note.get('id')
                note = note.get('note_card')
            display_title = note.get('display_title')
            user_name = note.get('user', {}).get('nick_name')
            if not user_name:
                user_name = note.get('user', {}).get('nickname')
            user_id = note.get('user', {}).get('user_id')
            liked_count = note.get('interact_info', {}).get('liked_count')
            cover = note.get('cover', {}).get('url')
            user_cover = note.get('user', {}).get('avatar')

            item = {
                'note_id': id,
                'display_title': display_title,
                'user_name': user_name,
                'user_id': user_id,
                'liked_count': liked_count,
                'cover': cover,
                'user_cover': user_cover
            }

            yield item

    # 解析搜索用户的全部信息
    @staticmethod
    def get_search_user(users: str):
        for user in users:
            id = user['id']
            fensi = user['fans']  # 粉丝
            note_count = user['note_count']  # 笔记
            image = user['image']  # 头像
            sub_title = user['sub_title']  # 小红书号
            name = user['name']
            try:
                update_time = user['update_time']  # 更新时间
            except Exception as e:

                update_time = ''

            item = {
                'id': id,
                'fensi': fensi,
                'note_count': note_count,
                'image': image,
                'sub_title': sub_title,
                'name': name,
                'update_time': update_time
            }
            yield item

    # 解析评论信息
    def get_comment(self, note_id, comments: str):
        list_1 = []
        for comment in comments:
            print(22222, comment)
            # 如果是回复别人的评论，则会有被回复的评论的信息，没有会显示主评论
            if comment.get('target_comment', {}):
                target_comment = {
                    'id': comment.get('target_comment', {}).get('id'),
                    'user': {
                        'id': comment.get('target_comment', {}).get('user_info', {}).get('user_id'),
                        'nickname': comment.get('target_comment', {}).get('user_info', {}).get('nickname')
                    }
                }
            else:
                target_comment = "root_comment"
            content = comment['content']
            id = comment['id']
            root_id = None
            note_id = note_id
            user_id = comment['user_info']['user_id']
            user_name = comment['user_info']['nickname']
            like_count = comment['like_count']
            at_users = comment.get('at_users')

            # 传递第一次的子评论
            if comment.get('sub_comments'):
                root_id = id
                yield from self.get_comment(note_id, comment['sub_comments'])
            # 模拟点击
            if comment.get('sub_comment_has_more') and root_id:
                sub_comment_cursor = comment.get('sub_comment_cursor')
                note_id = note_id
                root_comment_id = root_id
                print(sub_comment_cursor, note_id, root_comment_id)
                yield from self.more_comments(sub_comment_cursor, note_id, root_comment_id)

            item = {
                'target_comment': target_comment,
                'id': id,
                'user_id': user_id,
                'user_name': user_name,
                'content': content,
                'like_count': like_count,
                'at_users': at_users,
                '1': list_1

            }

            yield item

    # 模拟点击更多回复(有的时候会有多条回复需要模拟点击)
    def more_comments(self, cursor: str, note_id: str, root_id: str):
        continuations = [cursor]
        post_url = 'https://edith.xiaohongshu.com/api/sns/web/v2/comment/sub/page'
        while continuations:
            continuation = continuations.pop()
            params = {
                'note_id': note_id,
                'root_comment_id': root_id,
                'num': '10',
                'cursor': continuation,
            }
            # 发送请求
            self.post_url(url=post_url, data=params, x10=8, requestMethod='GET')
            print(self.response.text)
            has_more = self.response.json().get('data', {}).get('has_more')
            # 若还有更多评论，则继续点击
            if has_more:
                continuations.append(self.response.json()['data']['cursor'])
            data = self.response.json().get('data', {}).get('comments')
            if data:
                yield from self.get_comment(note_id, data)
            time.sleep(random.uniform(1, 5))

    # 指定用户id下载所有图片或者视频-
    def save(self, uid: str, option=0):
        if not os.path.exists('./upadate_'):
            os.mkdir('./upadate_')

        pics = self.get_all_urls(uid)
        print(pics)
        # 开辟线程池高效下载资源
        with ThreadPoolExecutor(max_workers=20) as executor:
            futures = []
            for pic in pics:
                print(pic)
                result = self.filter_mp4(pic)
                print(result)
                if result:
                    option = 1
                print(option)
                future = executor.submit(self.download, pic, option)
                futures.append(future)
            for future in futures:
                future.result()
            executor.shutdown()

    # 把session入库
    def save_data(self, data):
        # 获取表的名称
        sheet_name = list(data.keys())[0]
        os_mkdir_path = os.getcwd() + '/数据/'
        os.makedirs(os_mkdir_path, exist_ok=True)
        os_excel_path = os_mkdir_path + '数据.xlsx'

        if not os.path.exists(os_excel_path):
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            headers = ('链接', '作者', '用户id', '标题', '内容', '点赞数', '收藏数', '分享数', '评论数', '笔记类型', '性别', 'ip归属地', '关注数', '粉丝数', '点赞与收藏', '发布时间')
            worksheet.append(headers)
        else:
            workbook = load_workbook(os_excel_path)
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
            worksheet = workbook[sheet_name]

        # 从 data 字典中提取数据并添加到工作表
        row_data = data[sheet_name]
        if isinstance(row_data, list):
            worksheet.append(row_data)

        workbook.save(os_excel_path)


if __name__ == '__main__':
    #
    xhs = XHS('04006979f856730af5336cb3a6374b56361bc6')
    # items = xhs.search_topic('头像')
    #

    with open('ids(1).txt', 'r') as file:
        for line in file:
            try:
                items = xhs.get_note_detail(line.strip())
            except KeyError as e:
                logger.error(e)
                logger.error(line)
                break
            for i in items:
                logger.info(i)
            print(items)

    # aaa = xhs.get_user('5c09ea18000000000601df4d')
    # print(aaa)

    # aaa = xhs.get_usersearch('李四')
    # for i in aaa:
    #     print(i)

    # aaa = xhs.get_all_urls('65266506000000001a0211d2')
    # print(aaa)
    # xhs.save("651e0ed0000000001e02feb2")
